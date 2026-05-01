"""Build Layer 2 data for any state from official license roster.

Usage:
    python scripts/build_state_layer2.py --state PA
    python scripts/build_state_layer2.py --state NJ
    python scripts/build_state_layer2.py --state PA --sync-supabase

Each state has a config block defining:
- source file path
- header row number
- column mappings
- license type categories
- quota rules
"""

import argparse
import csv
import json
import math
import os
import re
import requests
import sys
from collections import Counter, defaultdict
from pathlib import Path

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))


# ============================================================
# STATE CONFIGURATIONS
# ============================================================

STATE_CONFIGS = {
    "NJ": {
        "state_fips": "34",
        "source_file": "data/seed/nj_retail_licensees_april2026.xlsx",
        "source_format": "xlsx",
        "header_row": 3,  # 1-indexed
        "data_start_row": 4,
        "source_url": "https://www.njoag.gov/wp-content/uploads/2026/04/RETAIL-LICENSE-REPORT-APRIL-2026.xlsx",
        "source_name": "NJ ABC Retail License Report April 2026",
        "column_map": {
            "license_number": "License Number",
            "license_type": "License Type",
            "establishment": "Establishment",
            "licensee": "Licensee",
            "city": "City",
            "address": "Premise Address",
            "county": None,  # derived from license number prefix
            "status": None,  # all active (pre-filtered)
            "effective_date": "Effective Date",
            "inactivity_date": "Inactivity Start Date",
        },
        "quota_rules": {
            "consumption_divisor": 3000,
            "distribution_divisor": 7500,
        },
        "license_categories": {
            "consumption": ["Plenary Retail Consumption License"],
            "broad_package": ["Plenary Retail Consumption License with Broad C"],
            "distribution": ["Plenary Retail Distribution License"],
            "limited_distribution": ["Limited Retail Distribution License"],
            "club": ["Club License"],
            "hotel": ["Hotel/Motel License"],
            "seasonal": ["Seasonal Retail Cons Lic 7/1-11/14 and 5/1-6/30"],
            "theater": ["Theater License"],
        },
        "county_from_license_prefix": True,
        "county_codes": {
            "01": "Atlantic", "02": "Bergen", "03": "Burlington", "04": "Camden",
            "05": "Cape May", "06": "Cumberland", "07": "Essex", "08": "Gloucester",
            "09": "Hudson", "10": "Hunterdon", "11": "Mercer", "12": "Middlesex",
            "13": "Monmouth", "14": "Morris", "15": "Ocean", "16": "Passaic",
            "17": "Salem", "18": "Somerset", "19": "Sussex", "20": "Union", "21": "Warren",
        },
    },
    "PA": {
        "state_fips": "42",
        "source_file": "data/seed/pa_license_list.csv",
        "source_format": "csv",
        "source_url": "https://plcbplus.pa.gov/pub/Default.aspx?PossePresentation=LicenseSearch",
        "source_name": "PA PLCB License Search Export April 2026",
        "column_map": {
            "license_number": "License Number",
            "license_type": "License Type",
            "establishment": "Premises",
            "licensee": "Licensee",
            "city": "Municipality",
            "address": "Premises Address",
            "county": "County",
            "status": "Status",
            "effective_date": "Last Issue Date",
            "expiration_date": "Expiration Date",
        },
        "active_status_value": "Active",
        "quota_rules": {
            "consumption_divisor": 3000,  # PA: 1 R license per 3,000 county pop
            "distribution_divisor": None,  # PA has no distribution quota (state stores)
            "quota_level": "county",  # PA quota is per COUNTY, not municipality
        },
        "license_categories": {
            "consumption": [
                "Restaurant (Liquor)",
                "Eating Place Retail Dispenser (Malt)",
                "Public Venue Restaurant",
                "Airport Restaurant (Liquor)",
                "Privately-Owned Public Golf Course Rest (Liquor)",
                "Municipal Golf Course (Liquor)",
                "Public Service (Liquor)",
                "Economic Development Restaurant (Liquor)",
                "Continuing Care Retirement Community (Liquor)",
            ],
            "hotel": ["Hotel (Liquor)"],
            "club": ["Club (Liquor)", "Catering Club (Liquor)"],
            "distributor": [
                "Distributor (Malt)",
                "Importing Distributor (Malt)",
            ],
            "manufacturer": [
                "Brewery", "Brewery Pub", "Brewery Storage",
                "Limited Winery", "Winery",
                "Limited Distillery", "Distillery",
            ],
            "performing_arts": ["Performing Arts Facility"],
        },
        "county_from_license_prefix": False,
        # PA is a CONTROL state: spirits+wine sold at state-run "Fine Wine & Good Spirits" stores
        # These stores are NOT in the license list (they're state-owned)
        "control_notes": "PA is control for spirits+wine. State-run Fine Wine & Good Spirits stores not in this data.",
    },
}


def load_source_data(config):
    """Load raw license data from source file."""
    source_path = PROJECT_ROOT / config["source_file"]
    fmt = config["source_format"]
    col_map = config["column_map"]

    rows = []

    if fmt == "csv":
        with open(source_path) as f:
            reader = csv.DictReader(f)
            for r in reader:
                # Filter to active only if status column exists
                if col_map.get("status"):
                    if r.get(col_map["status"], "") != config.get("active_status_value", "Active"):
                        continue
                rows.append({
                    "license_number": r.get(col_map["license_number"], ""),
                    "license_type": r.get(col_map["license_type"], ""),
                    "establishment": r.get(col_map["establishment"], ""),
                    "licensee": r.get(col_map["licensee"], ""),
                    "city": r.get(col_map["city"], ""),
                    "address": r.get(col_map["address"], ""),
                    "county": r.get(col_map["county"], "") if col_map.get("county") else "",
                    "effective_date": r.get(col_map.get("effective_date", ""), ""),
                    "expiration_date": r.get(col_map.get("expiration_date", ""), ""),
                })
    elif fmt == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(source_path)
        ws = wb.active
        header_row = config.get("header_row", 1)
        data_start = config.get("data_start_row", header_row + 1)
        headers = [cell.value for cell in ws[header_row]]

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            r = dict(zip(headers, row))
            if not r.get(col_map["license_number"]):
                continue
            rows.append({
                "license_number": str(r.get(col_map["license_number"], "") or "").strip(),
                "license_type": str(r.get(col_map["license_type"], "") or "").strip(),
                "establishment": str(r.get(col_map["establishment"], "") or "").strip(),
                "licensee": str(r.get(col_map["licensee"], "") or "").strip(),
                "city": str(r.get(col_map["city"], "") or "").strip(),
                "address": str(r.get(col_map.get("address", ""), "") or "").strip(),
                "county": "",
                "effective_date": str(r.get(col_map.get("effective_date", ""), "") or ""),
                "inactivity_date": str(r.get(col_map.get("inactivity_date", ""), "") or ""),
            })

    print(f"  Loaded {len(rows)} active licenses from {source_path.name}")
    return rows


def categorize_license(license_type, categories):
    """Map a license type string to a category."""
    for category, patterns in categories.items():
        for pattern in patterns:
            if pattern in license_type:
                return category
    return "other"


def get_census_population(state_fips):
    """Fetch municipality populations from Census API."""
    cache_path = PROJECT_ROOT / f"data/cache/census_pop_{state_fips}.json"
    if cache_path.exists():
        with open(cache_path) as f:
            return json.load(f)

    print(f"  Fetching Census population for state {state_fips}...")
    pop = {}

    # Places (cities, boroughs, towns)
    url = f"https://api.census.gov/data/2020/dec/pl?get=NAME,P1_001N&for=place:*&in=state:{state_fips}"
    resp = requests.get(url)
    if resp.status_code == 200:
        for row in resp.json()[1:]:
            name = row[0].split(",")[0].strip()
            for suffix in [" city", " borough", " town", " township", " village", " CDP"]:
                if name.lower().endswith(suffix):
                    name = name[:-len(suffix)].strip()
            pop[name.upper()] = int(row[1])

    # County subdivisions (MCDs/townships)
    url2 = f"https://api.census.gov/data/2020/dec/pl?get=NAME,P1_001N&for=county%20subdivision:*&in=state:{state_fips}"
    resp2 = requests.get(url2)
    if resp2.status_code == 200:
        for row in resp2.json()[1:]:
            name = row[0].split(",")[0].strip()
            for suffix in [" township", " town"]:
                if name.lower().endswith(suffix):
                    name = name[:-len(suffix)].strip()
            if name.upper() not in pop:
                pop[name.upper()] = int(row[1])

    # County populations (for PA quota which is county-based)
    url3 = f"https://api.census.gov/data/2020/dec/pl?get=NAME,P1_001N&for=county:*&in=state:{state_fips}"
    resp3 = requests.get(url3)
    if resp3.status_code == 200:
        for row in resp3.json()[1:]:
            name = row[0].split(",")[0].strip()
            name = name.replace(" County", "").strip()
            pop[f"COUNTY_{name.upper()}"] = int(row[1])

    cache_path.parent.mkdir(parents=True, exist_ok=True)
    with open(cache_path, "w") as f:
        json.dump(pop, f)

    print(f"  Cached {len(pop)} population entries")
    return pop


def build_municipality_summary(rows, config, pop_lookup):
    """Aggregate licenses by municipality."""
    state_fips = config["state_fips"]
    state_abbr = {"34": "NJ", "42": "PA"}.get(state_fips, "")
    categories = config["license_categories"]
    quota = config.get("quota_rules", {})

    # Aggregate
    muni_agg = defaultdict(lambda: {
        "county": "", "total": 0, "categories": Counter(), "establishments": []
    })

    for r in rows:
        city = r["city"].strip()
        if not city:
            continue

        md = muni_agg[city.upper()]
        md["total"] += 1
        if r["county"] and not md["county"]:
            md["county"] = r["county"].replace(" County", "").strip()
        cat = categorize_license(r["license_type"], categories)
        md["categories"][cat] += 1
        est = r.get("establishment", "")
        if est and est not in ("None", "Not Available", "NONE", "ABC POCKET"):
            md["establishments"].append(est)

    # Build output rows
    # Determine which categories exist for this state
    all_cats = set()
    for md in muni_agg.values():
        all_cats.update(md["categories"].keys())

    summary_rows = []
    for name in sorted(muni_agg.keys()):
        md = muni_agg[name]
        county = md["county"]

        # Population lookup
        pop = pop_lookup.get(name, 0)
        if not pop:
            for variant in [name + " CITY", name + " TOWNSHIP", "CITY OF " + name]:
                pop = pop_lookup.get(variant, 0)
                if pop:
                    break

        # Quota (PA is county-based, NJ is municipality-based)
        cons_div = quota.get("consumption_divisor")
        dist_div = quota.get("distribution_divisor")
        quota_level = quota.get("quota_level", "municipality")

        if quota_level == "county" and county:
            quota_pop = pop_lookup.get(f"COUNTY_{county.upper()}", 0)
        else:
            quota_pop = pop

        max_cons = max(1, math.floor(quota_pop / cons_div)) if cons_div and quota_pop else 0
        max_dist = max(1, math.floor(quota_pop / dist_div)) if dist_div and quota_pop else 0

        # Category counts
        consumption_count = sum(md["categories"].get(c, 0)
                                for c in ["consumption", "broad_package", "hotel", "seasonal",
                                           "theater", "performing_arts"])
        distribution_count = md["categories"].get("distribution", 0) + md["categories"].get("distributor", 0)
        limited_count = md["categories"].get("limited_distribution", 0)
        club_count = md["categories"].get("club", 0)
        hotel_count = md["categories"].get("hotel", 0)
        manufacturer_count = md["categories"].get("manufacturer", 0)

        has_consumption = consumption_count > 0
        has_distribution = distribution_count > 0
        has_limited = limited_count > 0
        has_club = club_count > 0
        has_hotel = hotel_count > 0
        has_manufacturer = manufacturer_count > 0

        # Quota notes
        cons_remaining = max_cons - consumption_count if max_cons else 0
        dist_remaining = max_dist - distribution_count if max_dist else 0
        over_cons = consumption_count > max_cons if max_cons else False
        over_dist = distribution_count > max_dist if max_dist else False

        notes_parts = []
        if max_cons:
            if over_cons:
                notes_parts.append(f"Over consumption: {consumption_count}/{max_cons} (grandfathered)")
            elif cons_remaining > 0:
                notes_parts.append(f"{cons_remaining} consumption slots available")
        if max_dist:
            if over_dist:
                notes_parts.append(f"Over distribution: {distribution_count}/{max_dist}")
            elif dist_remaining > 0:
                notes_parts.append(f"{dist_remaining} distribution slots available")
        quota_notes = "; ".join(notes_parts) if notes_parts else ("At capacity" if max_cons else "")

        summary_rows.append({
            "state_fips": state_fips,
            "state_abbr": state_abbr,
            "municipality_name": name.title(),
            "county_name": county,
            "total_active_licenses": md["total"],
            "has_consumption_license": str(has_consumption),
            "has_distribution_license": str(has_distribution),
            "has_limited_distribution": str(has_limited),
            "has_club_license": str(has_club),
            "has_hotel_license": str(has_hotel),
            "has_manufacturer": str(has_manufacturer),
            "consumption_count": consumption_count,
            "distribution_count": distribution_count,
            "limited_distribution_count": limited_count,
            "club_count": club_count,
            "hotel_count": hotel_count,
            "manufacturer_count": manufacturer_count,
            "grocery_can_sell_alcohol": str(has_distribution or has_limited or has_consumption),
            "population_2020": pop if pop else "",
            "max_consumption_quota": max_cons if max_cons else "",
            "consumption_slots_remaining": cons_remaining if max_cons else "",
            "max_distribution_quota": max_dist if max_dist else "",
            "distribution_slots_remaining": dist_remaining if max_dist else "",
            "is_over_consumption_quota": str(over_cons) if max_cons else "",
            "is_over_distribution_quota": str(over_dist) if max_dist else "",
            "quota_notes": quota_notes,
        })

    return summary_rows


def build_individual_records(rows, config):
    """Build individual license records for Supabase."""
    state_fips = config["state_fips"]
    state_abbr = {"34": "NJ", "42": "PA"}.get(state_fips, "")

    records = []
    for r in rows:
        city = r["city"].strip()
        if not city:
            continue

        records.append({
            "license_number": r["license_number"],
            "state_fips": state_fips,
            "municipality_name": city.title() if city else None,
            "county_name": r.get("county", "").replace(" County", "").strip() or None,
            "license_type": r["license_type"],
            "establishment_name": r["establishment"] if r["establishment"] not in ("None", "") else None,
            "licensee_name": r.get("licensee") if r.get("licensee") not in ("None", "") else None,
            "premise_address": r.get("address", "")[:200] if r.get("address") else None,
            "status": "Active",
            "effective_date": r.get("effective_date", "") or None,
        })

    return records


def write_csv(rows, output_path, fieldnames):
    """Write rows to CSV."""
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Wrote {len(rows)} rows to {output_path}")


def sync_to_supabase(summary_rows, individual_records, config):
    """Sync data to Supabase."""
    from supabase import create_client

    env_path = PROJECT_ROOT / ".env.local"
    url = key = ""
    with open(env_path) as f:
        for line in f:
            if line.startswith("SUPABASE_URL="):
                url = line.split("=", 1)[1].strip()
            if line.startswith("SUPABASE_KEY="):
                key = line.split("=", 1)[1].strip()

    client = create_client(url, key)
    S = "regulations_data"
    state_fips = config["state_fips"]

    # Clear existing data for this state
    client.schema(S).table("layer2_municipality_licenses").delete().eq("state_fips", state_fips).execute()
    client.schema(S).table("layer2_individual_licenses").delete().eq("state_fips", state_fips).execute()

    # Sync municipality summary
    for i in range(0, len(summary_rows), 400):
        batch = summary_rows[i:i + 400]
        supabase_rows = []
        for r in batch:
            def si(v):
                return int(v) if v and str(v).strip() and str(v).lstrip("-").isdigit() else 0

            def sb(v):
                return str(v).strip().lower() == "true" if v else False

            supabase_rows.append({
                "state_fips": r["state_fips"],
                "state_abbr": r["state_abbr"],
                "municipality_name": r["municipality_name"],
                "county_name": r.get("county_name") or None,
                "total_active_licenses": si(r.get("total_active_licenses")),
                "has_consumption_license": sb(r.get("has_consumption_license")),
                "has_distribution_license": sb(r.get("has_distribution_license")),
                "has_limited_distribution": sb(r.get("has_limited_distribution")),
                "has_club_license": sb(r.get("has_club_license")),
                "has_hotel_license": sb(r.get("has_hotel_license")),
                "consumption_count": si(r.get("consumption_count")),
                "distribution_count": si(r.get("distribution_count")),
                "limited_distribution_count": si(r.get("limited_distribution_count")),
                "club_count": si(r.get("club_count")),
                "hotel_count": si(r.get("hotel_count")),
                "grocery_can_sell_alcohol": sb(r.get("grocery_can_sell_alcohol")),
                "data_source": config["source_name"],
                "source_url": config["source_url"],
                "population_2020": si(r.get("population_2020")) or None,
                "max_consumption_quota": si(r.get("max_consumption_quota")) or None,
                "consumption_slots_remaining": int(r["consumption_slots_remaining"]) if str(
                    r.get("consumption_slots_remaining", "")).lstrip("-").isdigit() else None,
                "max_distribution_quota": si(r.get("max_distribution_quota")) or None,
                "distribution_slots_remaining": int(r["distribution_slots_remaining"]) if str(
                    r.get("distribution_slots_remaining", "")).lstrip("-").isdigit() else None,
                "is_over_consumption_quota": sb(r.get("is_over_consumption_quota")),
                "is_over_distribution_quota": sb(r.get("is_over_distribution_quota")),
                "quota_notes": r.get("quota_notes") or None,
            })

        client.schema(S).table("layer2_municipality_licenses").upsert(
            supabase_rows, on_conflict="state_fips,municipality_name"
        ).execute()

    print(f"  Synced {len(summary_rows)} municipality rows")

    # Deduplicate individual records globally
    seen_lic = set()
    deduped_records = []
    for r in individual_records:
        if r["license_number"] not in seen_lic:
            seen_lic.add(r["license_number"])
            deduped_records.append(r)

    # Sync individual licenses
    for i in range(0, len(deduped_records), 400):
        batch = deduped_records[i:i + 400]
        client.schema(S).table("layer2_individual_licenses").upsert(
            batch, on_conflict="license_number"
        ).execute()
        if (i // 400) % 5 == 0:
            print(f"  Individual: {i + len(batch)}/{len(deduped_records)}...")

    print(f"  Synced {len(individual_records)} individual licenses")


def main():
    parser = argparse.ArgumentParser(description="Build Layer 2 data for a state")
    parser.add_argument("--state", required=True, choices=STATE_CONFIGS.keys(),
                        help="State abbreviation (NJ, PA)")
    parser.add_argument("--sync-supabase", action="store_true",
                        help="Sync results to Supabase")
    args = parser.parse_args()

    config = STATE_CONFIGS[args.state]
    state_fips = config["state_fips"]
    state_abbr = args.state

    print(f"\n{'=' * 60}")
    print(f"Building Layer 2 for {state_abbr} (FIPS {state_fips})")
    print(f"{'=' * 60}")

    # 1. Load source data
    print("\n1. Loading source data...")
    rows = load_source_data(config)

    # 2. Get Census population
    print("\n2. Fetching Census population...")
    pop = get_census_population(state_fips)

    # 3. Build municipality summary
    print("\n3. Building municipality summary...")
    summary = build_municipality_summary(rows, config, pop)

    # 4. Build individual records
    print("\n4. Building individual license records...")
    individual = build_individual_records(rows, config)

    # 5. Write CSVs
    print("\n5. Writing output files...")
    summary_path = PROJECT_ROOT / f"data/seed/{state_abbr.lower()}_municipality_license_summary.csv"
    fieldnames = list(summary[0].keys()) if summary else []
    write_csv(summary, summary_path, fieldnames)

    # 6. Sync to Supabase
    if args.sync_supabase:
        print("\n6. Syncing to Supabase...")
        sync_to_supabase(summary, individual, config)

    # 7. Summary stats
    total_lic = sum(int(r.get("total_active_licenses", 0)) for r in summary)
    with_lic = sum(1 for r in summary if int(r.get("total_active_licenses", 0)) > 0)

    print(f"\n{'=' * 60}")
    print(f"DONE — {state_abbr} Layer 2")
    print(f"{'=' * 60}")
    print(f"  Municipalities: {len(summary)} ({with_lic} with licenses)")
    print(f"  Total licenses: {total_lic}")
    print(f"  Individual records: {len(individual)}")
    print(f"  Source: {config['source_name']}")
    print(f"  Output: {summary_path}")

    # Key chains search
    chains = ["SHEETZ", "WAWA", "GIANT", "WEGMANS", "COSTCO", "WALMART", "7-ELEVEN",
              "TRADER JOE", "WHOLE FOODS", "SHOPRITE", "WEIS", "ALDI", "TARGET"]
    print(f"\n  Chain store licenses:")
    for chain in chains:
        matches = [r for r in rows if chain in str(r.get("establishment", "")).upper()
                   or chain in str(r.get("licensee", "")).upper()]
        if matches:
            print(f"    {chain}: {len(matches)} licenses")


if __name__ == "__main__":
    main()
